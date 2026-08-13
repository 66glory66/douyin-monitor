"""关键词雷达运行编排：搜索 → 评论/回复 → 评分 → 拆解 → 导出。"""
import asyncio
import csv
import json
import logging
import time
from datetime import datetime
from pathlib import Path

import requests

from comment_spider import CommentSpider
from config_manager import load_config
from keyword_spider import KeywordSpider
from lead_scoring import analyze_content, classify_intent, interaction_score, score_work
from radar_db import (
    add_feedback,
    add_work_snapshot,
    get_radar_conversion_summary,
    get_radar_work,
    init_radar_db,
    list_radar_comments,
    list_radar_works,
    upsert_content_analysis,
    upsert_keyword,
    upsert_radar_comment,
    upsert_work,
    update_work_score,
)

logger = logging.getLogger(__name__)


class RadarRunner:
    def __init__(self, config: dict | None = None, headless: bool | None = None):
        self.config = config or load_config()
        self.radar_cfg = self.config.get("radar", {})
        self.headless = headless

    async def run(
        self,
        keywords: list[str] | None = None,
        days: int | None = None,
        limit: int | None = None,
        fetch_comments: bool = True,
    ) -> dict:
        init_radar_db()
        keywords = [k.strip() for k in (keywords or self.radar_cfg.get("keywords", [])) if k.strip()]
        days = int(days or self.radar_cfg.get("days", 7))
        limit = int(limit or self.radar_cfg.get("max_results_per_keyword", 20))
        if not keywords:
            raise ValueError("没有配置关键词，请在 config.yaml 的 radar.keywords 中添加")

        searcher = KeywordSpider(headless=self.headless)
        discovered = 0
        work_ids: set[int] = set()
        errors: list[str] = []
        for keyword in keywords:
            upsert_keyword(keyword)
            try:
                works = await searcher.search(keyword, days=days, limit=limit)
            except Exception as exc:
                logger.error("关键词 %s 搜索失败: %s", keyword, exc)
                errors.append(f"{keyword}: {exc}")
                continue
            for work in works:
                work_id = upsert_work(work)
                add_work_snapshot(work_id, work)
                work_ids.add(work_id)
                discovered += 1
            logger.info("关键词 %s：发现 %d 条近%d日作品", keyword, len(works), days)

        works = [w for w in list_radar_works(limit=max(200, len(work_ids) + 20), days=days)
                 if w["id"] in work_ids]
        if fetch_comments and works:
            top_n = int(self.radar_cfg.get("top_works_for_comments", 10))
            works_for_comments = sorted(
                works, key=interaction_score, reverse=True
            )[:top_n]
            await self._fetch_and_score_comments(works_for_comments, days=days)

        works = [w for w in list_radar_works(limit=max(200, len(work_ids) + 20), days=days)
                 if w["id"] in work_ids]
        self._score_and_analyze(works)
        works = [w for w in list_radar_works(limit=max(200, len(work_ids) + 20), days=days)
                 if w["id"] in work_ids]
        files = self.export(works=works, days=days)
        feishu = self.send_feishu(works, files)
        return {
            "keywords": keywords,
            "days": days,
            "discovered": discovered,
            "works": len(works),
            "files": files,
            "feishu": feishu,
            "errors": errors,
        }

    async def _fetch_and_score_comments(self, works: list[dict], days: int = 7) -> None:
        cfg = self.radar_cfg
        pages = int(cfg.get("comment_pages", 5))
        max_total = int(cfg.get("max_comments_per_work", 100))
        for index, work in enumerate(works, start=1):
            vid = work.get("platform_work_id")
            if not vid:
                continue
            try:
                spider = CommentSpider(headless=self.headless)
                comments = await spider.fetch_comments(
                    vid,
                    max_pages=pages,
                    max_total=max_total,
                    apply_filter=False,
                    include_replies=True,
                )
            except Exception as exc:
                logger.warning("作品 %s 评论抓取失败: %s", vid, exc)
                continue
            since_ts = int(time.time()) - max(1, days) * 86400
            comments = [
                c for c in comments
                if not c.get("create_time") or int(c.get("create_time", 0)) >= since_ts
            ]
            for comment in comments:
                intent = classify_intent(comment.get("text", ""))
                comment.update(
                    intent_label=intent["label"],
                    intent_score=intent["score"],
                    intent_reason=intent["reason"],
                )
                upsert_radar_comment(work["id"], comment)
            logger.info("评论 %d/%d：%s，采集 %d 条（含可见回复）", index, len(works), vid, len(comments))
            if index < len(works):
                await asyncio.sleep(1.5)

    def _score_and_analyze(self, works: list[dict]) -> None:
        cfg = self.radar_cfg
        ai_config = {
            "enabled": bool(cfg.get("ai_enabled")),
            "base_url": cfg.get("ai_base_url", ""),
            "model": cfg.get("ai_model", ""),
        }
        for work in works:
            comments = list_radar_comments(work_id=work["id"], limit=1000)
            peers = self._peer_interactions(work)
            score = score_work(work, comments, peers)
            update_work_score(work["id"], score)
            analysis = analyze_content(work, transcript=self._transcript(work["platform_work_id"]), ai_config=ai_config)
            upsert_content_analysis(work["id"], analysis)

    @staticmethod
    def _peer_interactions(work: dict) -> list[int]:
        from db import get_db

        author = work.get("author_sec_uid") or work.get("author_name")
        if not author:
            return []
        with get_db() as db:
            rows = db.execute(
                "SELECT like_count,comment_count,share_count,collect_count "
                "FROM radar_works WHERE (author_sec_uid=? OR author_name=?) "
                "AND id != ? ORDER BY create_time DESC LIMIT 30",
                (author, author, work["id"]),
            ).fetchall()
        return [interaction_score(dict(row)) for row in rows]

    @staticmethod
    def _transcript(platform_work_id: str) -> str:
        from db import get_db

        with get_db() as db:
            row = db.execute(
                "SELECT t.full_text FROM transcripts t JOIN videos v ON v.id=t.video_id "
                "WHERE v.video_id=? LIMIT 1",
                (platform_work_id,),
            ).fetchone()
        return row["full_text"] if row and row["full_text"] else ""

    def export(self, works: list[dict], days: int) -> dict:
        out_dir = Path(self.radar_cfg.get("output_dir", "./exports/radar"))
        if not out_dir.is_absolute():
            out_dir = Path(__file__).parent / out_dir
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        selected_ids = {w["id"] for w in works}
        comments = [
            c for c in list_radar_comments(limit=5000)
            if c.get("work_id") in selected_ids
        ]
        conversion = get_radar_conversion_summary(days)
        payload = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "days": days,
            "works": works,
            "comments": comments,
            "conversion": conversion,
        }
        json_path = out_dir / f"radar_report_{stamp}.json"
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

        csv_path = out_dir / f"radar_works_{stamp}.csv"
        with csv_path.open("w", newline="", encoding="utf-8-sig") as fp:
            writer = csv.writer(fp)
            writer.writerow([
                "作品ID", "标题", "作者", "发布时间", "点赞", "评论", "分享", "收藏",
                "Content Boom分", "高意向评论", "意向密度", "账号相对倍数", "内容路径",
                "可复制性", "链接", "评分说明",
            ])
            for row in works:
                writer.writerow([
                    row.get("platform_work_id"), row.get("title"), row.get("author_name"),
                    _fmt_ts(row.get("create_time")), row.get("like_count", 0),
                    row.get("comment_count", 0), row.get("share_count", 0), row.get("collect_count", 0),
                    row.get("boom_score", 0), row.get("high_intent_count", 0),
                    row.get("intent_density", 0), row.get("relative_ratio", 0),
                    row.get("content_path", ""), row.get("replicability", 0), row.get("url", ""),
                    row.get("scoring_reason", ""),
                ])

        comment_csv_path = out_dir / f"radar_leads_{stamp}.csv"
        with comment_csv_path.open("w", newline="", encoding="utf-8-sig") as fp:
            writer = csv.writer(fp)
            writer.writerow(["作品ID", "作品标题", "评论ID", "回复", "评论", "用户", "点赞", "意向标签", "意向分", "原因", "人工状态"])
            for row in comments:
                writer.writerow([
                    row.get("platform_work_id"), row.get("title"), row.get("comment_id"),
                    "是" if row.get("is_reply") else "否", row.get("text"), row.get("user_name"),
                    row.get("digg_count", 0), row.get("intent_label"), row.get("intent_score", 0),
                    row.get("intent_reason"), row.get("feedback_status", ""),
                ])

        xlsx_path = self._export_xlsx(out_dir, stamp, works, comments, conversion)
        return {
            "json": str(json_path),
            "works_csv": str(csv_path),
            "leads_csv": str(comment_csv_path),
            "xlsx": str(xlsx_path) if xlsx_path else "",
        }

    @staticmethod
    def _export_xlsx(
        out_dir: Path,
        stamp: str,
        works: list[dict],
        comments: list[dict],
        conversion: dict,
    ) -> Path | None:
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.warning("openpyxl 未安装，跳过 Excel xlsx；CSV 仍已输出")
            return None
        path = out_dir / f"radar_export_{stamp}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "作品评分"
        ws.append(["作品ID", "标题", "作者", "发布时间", "Boom分", "高意向", "意向密度", "内容路径", "可复制性", "咨询", "加微信", "成交", "金额", "链接"])
        for row in works:
            ws.append([
                row.get("platform_work_id"), row.get("title"), row.get("author_name"),
                _fmt_ts(row.get("create_time")), row.get("boom_score", 0),
                row.get("high_intent_count", 0), row.get("intent_density", 0),
                row.get("content_path", ""), row.get("replicability", 0),
                row.get("consulted_count", 0), row.get("wechat_count", 0),
                row.get("purchased_count", 0), row.get("revenue", 0), row.get("url", ""),
            ])
        lead_ws = wb.create_sheet("高意向评论")
        lead_ws.append(["作品ID", "作品标题", "评论", "用户", "意向标签", "意向分", "原因", "人工状态"])
        for row in comments:
            if row.get("intent_label") not in {"high", "medium"}:
                continue
            lead_ws.append([
                row.get("platform_work_id"), row.get("title"), row.get("text"), row.get("user_name"),
                row.get("intent_label"), row.get("intent_score", 0), row.get("intent_reason"),
                row.get("feedback_status", ""),
            ])
        feedback_ws = wb.create_sheet("人工回填")
        feedback_ws.append(["作品ID", "评论ID", "是否咨询", "是否加微信", "是否成交", "产品", "金额", "备注"])
        for row in comments:
            feedback_ws.append([row.get("platform_work_id"), row.get("id"), "否", "否", "否", "", 0, ""])
        summary_ws = wb.create_sheet("转化统计")
        summary_ws.append(["统计项", "数值"])
        for key, label in (
            ("works", "作品数"), ("comments", "评论数"), ("high_intent_comments", "高意向评论"),
            ("consulted", "已咨询"), ("added_wechat", "已加微信"),
            ("purchased", "已成交"), ("revenue", "成交金额"),
            ("consult_rate", "高意向→咨询率"), ("wechat_rate", "咨询→加微信率"),
            ("purchase_rate", "加微信→成交率"),
        ):
            summary_ws.append([label, conversion.get(key, 0)])
        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for col in sheet.columns:
                width = min(42, max(12, max(len(str(cell.value or "")) for cell in col) + 2))
                sheet.column_dimensions[col[0].column_letter].width = width
        wb.save(path)
        return path

    def send_feishu(self, works: list[dict], files: dict) -> dict:
        webhook = self.radar_cfg.get("feishu_webhook", "")
        if not webhook:
            return {"sent": False, "reason": "未配置 radar.feishu_webhook"}
        top = works[:5]
        lines = [f"抖音 AI 工具会员雷达（近{self.radar_cfg.get('days', 7)}日）", f"作品数：{len(works)}"]
        for idx, row in enumerate(top, start=1):
            lines.append(
                f"{idx}. {row.get('title', '')[:45]} | Boom {row.get('boom_score', 0)} | "
                f"高意向 {row.get('high_intent_count', 0)} | {row.get('url', '')}"
            )
        lines.append(f"Excel：{files.get('xlsx') or files.get('works_csv')}")
        try:
            response = requests.post(
                webhook,
                json={"msg_type": "text", "content": {"text": "\n".join(lines)}},
                timeout=15,
            )
            response.raise_for_status()
            return {"sent": True}
        except Exception as exc:
            logger.warning("飞书 webhook 发送失败：%s", exc)
            return {"sent": False, "reason": str(exc)}


def _fmt_ts(value) -> str:
    try:
        return datetime.fromtimestamp(int(value)).strftime("%Y-%m-%d %H:%M") if value else ""
    except (TypeError, ValueError, OSError):
        return ""
